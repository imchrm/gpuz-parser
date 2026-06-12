from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.models import Company

COLUMNS: list[tuple[str, str]] = [
    ("name", "Название"),
    ("alt_names", "Доп. названия"),
    ("phones", "Телефоны"),
    ("website", "Сайт"),
    ("telegram", "Telegram"),
    ("city", "Город"),
    ("region", "Регион"),
    ("district", "Район"),
    ("street", "Улица"),
    ("building", "Дом/офис"),
    ("postal_code", "Индекс"),
    ("landmarks", "Ориентиры"),
    ("activity_types", "Виды деятельности"),
    ("inn", "ИНН"),
    ("last_updated", "Обновлено"),
    ("url", "URL"),
]


def _get_value(company: Company, field: str) -> str:
    value = getattr(company, field, None)
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(v) for v in value)
    return str(value)


def export_excel(companies: list[Company], output_path: str) -> str:
    path = Path(output_path).with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    headers = [header for _, header in COLUMNS]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    ws.freeze_panes = "A2"

    for company in companies:
        ws.append([_get_value(company, field) for field, _ in COLUMNS])

    for col_idx, _ in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(path)
    return str(path)
